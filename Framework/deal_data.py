"""

@author: lingw
@email: gw.lin@hzgosun.com
@file: deal_data.py
@time: 2020-4-23 上午 10:51

首先，如果商品的描述和图片与实物严重不符，具有欺诈性质甚至涉嫌诈骗，这时买家可以选择报警，或者依据《消费者权益保护法》要求卖家双倍赔偿货款。
其次，如果商品的描述和图片与实物不符但未到欺诈的程度，买家可以依据《合同法》要求解除合同并返还货款也就是退货。
最后，如果商品的描述和图片与实物只是略微不符，例如衣服显示色差等瑕疵，尚不构成根本违约，买家就不能要求卖家退货，只能下次网购的时候多加注意。

"""


import openpyxl
import requests
import os

Test_Case_Path = os.path.abspath(os.path.join(os.getcwd(), "..\\Test_Case"))

def get_excel_text(excle_path):
    # 获取excel文件路径
    exccle_file = Test_Case_Path + excle_path
    wb = openpyxl.load_workbook(exccle_file)
    sheets = wb.sheetnames
    # 获取excel第一个sheet
    ws = wb[sheets[0]]
    # 共多少行
    rows = ws.max_row
    # 共多少列
    cols = ws.max_column

    if rows < 2:
        return "该excel除了表头还有别的吗"
    else:
        list = []
        for j in range(2, rows + 1):
            j = 3
            url = ws.cell(j, 4).value
            headers = ws.cell(j, 5).value
            request_meathod = ws.cell(j, 6).value
            data = ws.cell(j, 7).value
            single_request = get_requests(request_method=request_meathod, headers=headers, url=url, data=data)
            list.append(single_request)
        return list
            # print(single_request)

def get_requests(request_method, headers, url, data):
    if request_method == 'get' or request_method == 'GET':
        request = requests.get(url=url, params=data).json()
        return request
    elif request_method == 'post' or request_method == 'POST':
        request = requests.post(url=url, headers=headers, data=data).json()
        return request
    elif request_method == 'put' or request_method == 'PUT':
        request = requests.put(url=url, headers=headers, data=data).json()
        return request
    elif request_method == 'delete' or request_method == 'DELETE':
        request = requests.delete(url=url).json()
        return request
    else:
        return "方法不在get,post,put,delete范围内"



if __name__ == "__main__":
    get_excel_text("/test.xlsx")