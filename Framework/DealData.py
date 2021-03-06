"""

@author: lingw
@email: gw.lin@hzgosun.com
@file: DealData.py
@time: 2020-4-23 上午 10:51

首先，如果商品的描述和图片与实物严重不符，具有欺诈性质甚至涉嫌诈骗，这时买家可以选择报警，或者依据《消费者权益保护法》要求卖家双倍赔偿货款。
其次，如果商品的描述和图片与实物不符但未到欺诈的程度，买家可以依据《合同法》要求解除合同并返还货款也就是退货。
最后，如果商品的描述和图片与实物只是略微不符，例如衣服显示色差等瑕疵，尚不构成根本违约，买家就不能要求卖家退货，只能下次网购的时候多加注意。

"""


import openpyxl


from setting import dicfile

Test_Case_Path = dicfile() +  "\\TestCase"

def get_excel_text(excle_path):
    # 获取excel文件路径
    excle_file = Test_Case_Path + excle_path
    print(excle_file)
    wb = openpyxl.load_workbook(excle_file)
    sheets = wb.sheetnames
    list = []
    # 获取excel所有sheet
    for sheet in sheets:
        ws = wb[sheet]
        # 共多少行
        rows = ws.max_row
        # 共多少列
        cols = ws.max_column
        if rows < 2:
            print("{}下的sheet:{}除了表头还有别的吗,你不会连表头都没有吧".format(excle_path,sheet))
        else:
            for row in range(2, rows + 1):
                dict = {}
                for col in range(1, cols+1):
                    dict[ws.cell(1, col).value] = ws.cell(row, col).value
                list.append(dict)
    return list
    # print(single_request)



if __name__ == "__main__":
    get_excel_text("/test.xlsx")